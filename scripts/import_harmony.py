"""
Accurate Harmony Auto / BYD Melbourne import.

Status + Vehicle Location combinations encountered in the source spreadsheet
have been audited and mapped to platform stages explicitly. See STAGE_RULES.

Run:
    python3 /app/backend/scripts/import_harmony.py [path-to-xlsx]
"""
import sys
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402
from dotenv import load_dotenv  # noqa: E402
from pymongo import MongoClient  # noqa: E402

load_dotenv(Path(__file__).resolve().parents[1] / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ.get('DB_NAME', 'test_database')

# Vehicle Locations that mean the car is physically at our delivery centre.
DEALERSHIP_LOCATIONS = {
    'FAIRFIELD',
    'NUNAWADING',
    'CAROLINE SPRINGS',
    'DERRIMUT HOLDING YARD',
    'DEALERSHIP',
}

# Vehicle Locations that mean the car is in transit / not yet at the dealership.
TRANSIT_LOCATIONS = {
    'CLAYTON HOLDING YARD',  # BYD AU regional distribution
    'ON WATER',              # in shipping
    'DISPATCH',              # being dispatched from supplier
    'IN TRANSIT',
}

# Locations that signal the deal is effectively done.
DELIVERED_LOCATIONS = {'DELIVERED'}


def normalise_phone(raw):
    if not raw:
        return ''
    digits = re.sub(r'[^\d+]', '', str(raw))
    if digits.startswith('+'):
        return digits
    if digits.startswith('61') and len(digits) == 11:
        return '+' + digits
    if digits.startswith('04') and len(digits) == 10:
        return '+61' + digits[1:]
    if digits.startswith('4') and len(digits) == 9:
        return '+61' + digits
    return digits


def to_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str) and v.strip():
        s = v.strip()
        if s.upper() in ('TBA', 'PENDING', 'N/A', 'NA'):
            return None
        m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', s)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
        if m:
            return s
    return None


def title_case_name(s):
    if not s:
        return ''
    s = str(s).strip()
    if re.search(r'\b(PTY|LTD|GROUP|HOLDING|TRUST)\b', s, re.I):
        return s
    return ' '.join(p.capitalize() for p in re.split(r'\s+', s))


def derive_stage_and_flags(get):
    """Return (stage, arrived, vehicle_at, skip_reason).

    Mapping rules audited against May-26 dataset:
    - status DELIVERED OR Actual Delivery date  -> Delivered, arrived
    - status CANCELLED OR blank                 -> SKIP
    - veh_loc DELIVERED                         -> Delivered, arrived
    - veh_loc Fairfield/Nunawading/Carol-Spr/Derrimut -> dealership; PDI date determines:
            PDI date set                        -> Ready for Pickup
            otherwise                           -> Pre-Delivery Inspection
        arrived=True
    - veh_loc Clayton Holding Yard / On Water / Dispatch -> In Transit, not arrived
    - veh_loc No Allocation OR status NOT AVAILABLE       -> Scheduled, not arrived
    - everything else                                     -> Scheduled, not arrived
    """
    status = (get('Status') or '').strip().upper()
    veh_loc = (get('Vehicle Location') or '').strip().upper()
    pdi_done = bool(get('PDI Completion Date'))
    actual = bool(get('Actual Delivery'))

    if status == 'CANCELLED':
        return None, False, veh_loc, 'cancelled'
    if status == '' and not get("Driver's Name") and not get('Client'):
        return None, False, veh_loc, 'blank'

    # Already delivered
    if actual or status == 'DELIVERED' or veh_loc in DELIVERED_LOCATIONS:
        return 'Delivered', True, veh_loc, None

    # At our dealership
    if veh_loc in DEALERSHIP_LOCATIONS:
        return ('Ready for Pickup' if pdi_done else 'Pre-Delivery Inspection'), True, veh_loc, None

    # In transit
    if veh_loc in TRANSIT_LOCATIONS:
        return 'In Transit', False, veh_loc, None

    # Awaiting allocation / no vehicle yet
    if status == 'NOT AVAILABLE' or veh_loc == 'NO ALLOCATION' or veh_loc == '':
        return 'Scheduled', False, veh_loc, None

    # Fallback
    return 'Scheduled', False, veh_loc, None


def derive_contact_status(get, stage):
    if stage == 'Delivered':
        return 'Contacted'
    has_followup = bool((get('Follow Up Comments') or '').strip())
    has_sales = bool((get('Sales Comments') or '').strip())
    if has_followup:
        return 'Contacted'
    if has_sales:
        return 'Awaiting Reply'
    return 'Not Contacted'


def build_vehicle(get):
    parts = []
    veh_type = (get('VEH. Type') or '').strip()
    model = (get('VEH. Model') or '').strip()
    variant = (get('VEH. Variant') or '').strip()
    colour = (get('VEH. Colour') or '').strip()
    if veh_type:
        parts.append(f"BYD {veh_type.title()}")
    if model and model.lower() not in (veh_type.lower(),):
        parts.append(model.title())
    if variant:
        parts.append(variant.title())
    label = ' '.join(parts).strip()
    if colour:
        label = f"{label} \u00b7 {colour.title()}" if label else colour.title()
    return label or 'BYD Vehicle'


def split_addons(get):
    out = []
    for key in ('ACCESSORIES', 'A/Mkt External Accessories', 'Aftermarket Products'):
        v = get(key)
        if not v:
            continue
        for line in re.split(r'[\n;,]+', str(v)):
            line = line.strip(' -*')
            if line and len(line) > 2:
                out.append(line[:120])
    seen = set()
    deduped = []
    for a in out:
        k = a.lower()
        if k not in seen:
            seen.add(k); deduped.append(a)
    return deduped


def aftermarket_text(get):
    parts = []
    for key in ('A/M Comments', 'Stock Control Notes'):
        v = get(key)
        if v and str(v).strip():
            parts.append(f"[{key}] {str(v).strip()}")
    return '\n'.join(parts) or None


def comments_blocks(get, vehicle_at):
    """Return a list of comment dicts to attach to the client."""
    out = []
    sources = [
        ('Pre-Delivery Comments', 'Pre-Delivery'),
        ('Sales Comments', 'Sales'),
        ('Follow Up Comments', 'Follow-up'),
    ]
    for col_name, label in sources:
        body = get(col_name)
        if not body:
            continue
        clean = str(body).strip()
        if not clean or clean.upper() in ('PDI COMPLETED',):
            continue
        out.append({
            'id': uuid.uuid4().hex,
            'author_id': None,
            'author_name': f'Harmony Auto · {label}',
            'body': clean,
            'created_at': datetime.now(timezone.utc),
        })
    if vehicle_at:
        out.append({
            'id': uuid.uuid4().hex,
            'author_id': None,
            'author_name': 'System',
            'body': f'Vehicle location at import: {vehicle_at}',
            'created_at': datetime.now(timezone.utc),
        })
    return out


def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    # Wipe previously imported Harmony rows so we start clean
    wiped = db.clients.delete_many({'imported_from': 'harmony-xlsx'})
    print(f"Wiped previous Harmony import: {wiped.deleted_count}")

    user_map = {u.get('name', '').lower(): u for u in db.users.find({'active': True})}

    counts = {
        'inserted': 0,
        'skipped_cancelled': 0,
        'skipped_blank': 0,
        'by_stage': {},
        'arrived': 0,
        'unassigned': 0,
    }

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue

        def get(k, _row=r):
            i = col(k)
            return _row[i] if i is not None and i < len(_row) else None

        name_raw = get("Driver's Name") or get('Client') or get('Company Name')
        if not name_raw:
            counts['skipped_blank'] += 1
            continue

        stage, arrived, vehicle_at, skip = derive_stage_and_flags(get)
        if skip:
            counts[f'skipped_{skip}'] = counts.get(f'skipped_{skip}', 0) + 1
            continue

        order_no = (str(get('Order No.') or '').strip()) or None
        phone = normalise_phone(get('Phone Number'))
        email = (str(get('Email') or '').strip().lower()) or None
        if email == '':
            email = None

        delivery_date = to_iso_date(get('Estimated Delivery')) or to_iso_date(get('Actual Delivery'))
        actual_delivery = to_iso_date(get('Actual Delivery'))
        order_date = to_iso_date(get('Order Date'))

        sales_person = (str(get('Sales Person') or '').strip()) or None
        delivery_consultant = (str(get('Delivery Consultant') or '').strip()) or None

        assigned_agent_id = None
        if delivery_consultant:
            u = user_map.get(delivery_consultant.lower())
            if u:
                assigned_agent_id = u['id']

        addons = split_addons(get)
        aftermarket = aftermarket_text(get)
        comments = comments_blocks(get, vehicle_at if not arrived else None)
        contact_status = derive_contact_status(get, stage)

        suburb = (str(get('Suburb') or '').strip().title()) or None
        location = None
        if suburb:
            location = suburb if 'VIC' in suburb.upper() else f"{suburb}, VIC"

        now = datetime.now(timezone.utc)
        last_contacted = now if contact_status in ('Contacted', 'Awaiting Reply') else None
        if stage == 'Delivered' and actual_delivery:
            try:
                last_contacted = datetime.fromisoformat(f'{actual_delivery}T12:00:00+00:00')
            except Exception:
                last_contacted = now

        client_doc = {
            'id': uuid.uuid4().hex,
            'name': title_case_name(name_raw),
            'phone': phone or '',
            'email': email,
            'vehicle': build_vehicle(get),
            'rego': (str(get('Rego No.') or '').strip()) or None,
            'vin': (str(get('Vin No.') or '').strip()) or None,
            'delivery_date': delivery_date,
            'stage': stage,
            'salesperson': sales_person,
            'notes': None,
            'address': None,
            'location': location,
            'deal_type': 'Retail',
            'vy_order_id': order_no,
            'vy_stock_id': (str(get('Stock No.') or '').strip()) or None,
            'stripe_customer_id': None,
            'arrived': arrived,
            'arrived_at': now if arrived else None,
            'contact_status': contact_status,
            'last_contacted_at': last_contacted,
            'assigned_agent_id': assigned_agent_id,
            'accessories': [],
            'aftermarket_notes': aftermarket,
            'addons': addons,
            'imported_from': 'harmony-xlsx',
            'imported_at': now,
            'comments': comments,
            'created_at': datetime.fromisoformat(f'{order_date}T00:00:00+00:00') if order_date else now,
            'updated_at': now,
        }
        db.clients.insert_one(client_doc)
        counts['inserted'] += 1
        counts['by_stage'][stage] = counts['by_stage'].get(stage, 0) + 1
        if arrived:
            counts['arrived'] += 1
        if not assigned_agent_id:
            counts['unassigned'] += 1

    print('\n=== Import summary ===')
    print(f"Inserted:           {counts['inserted']}")
    print(f"Skipped (cancelled): {counts.get('skipped_cancelled', 0)}")
    print(f"Skipped (blank):    {counts.get('skipped_blank', 0)}")
    print('\nBy stage:')
    for s, n in sorted(counts['by_stage'].items(), key=lambda x: -x[1]):
        print(f"  {s:30s} {n}")
    print(f"\nArrived (at dealership): {counts['arrived']}")
    print(f"Unassigned (no agent):   {counts['unassigned']}")
    print(f"Total clients in DB now: {db.clients.count_documents({})}")


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/tmp/harmony2.xlsx'
    main(path)
