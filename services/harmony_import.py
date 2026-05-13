"""
Harmony Auto Excel import service.

Reusable parser used by both the CLI script and the admin upload endpoint.
Accepts an xlsx file path or BytesIO, returns (summary_dict, replace_mode).
"""
import io
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from pymongo import MongoClient

DEALERSHIP_LOCATIONS = {
    'FAIRFIELD', 'NUNAWADING', 'CAROLINE SPRINGS', 'DERRIMUT HOLDING YARD', 'DEALERSHIP',
}
TRANSIT_LOCATIONS = {
    'CLAYTON HOLDING YARD', 'ON WATER', 'DISPATCH', 'IN TRANSIT',
}
DELIVERED_LOCATIONS = {'DELIVERED'}


def normalise_phone(raw):
    if not raw:
        return ''
    digits = re.sub(r'[^\d+]', '', str(raw))
    if digits.startswith('+'):
        return digits
    if digits.startswith('61') and len(digits) == 11:
        return '+' + digits
    if digits.startswith('04') and len(digits) == 10:
        return '+61' + digits[1:]
    if digits.startswith('4') and len(digits) == 9:
        return '+61' + digits
    return digits


def to_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str) and v.strip():
        s = v.strip()
        if s.upper() in ('TBA', 'PENDING', 'N/A', 'NA'):
            return None
        m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', s)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
        if m:
            return s
    return None


def title_case_name(s):
    if not s:
        return ''
    s = str(s).strip()
    if re.search(r'\b(PTY|LTD|GROUP|HOLDING|TRUST)\b', s, re.I):
        return s
    return ' '.join(p.capitalize() for p in re.split(r'\s+', s))


def derive_stage(get):
    status = (get('Status') or '').strip().upper()
    veh_loc = (get('Vehicle Location') or '').strip().upper()
    pdi_done = bool(get('PDI Completion Date'))
    actual = bool(get('Actual Delivery'))

    if status == 'CANCELLED':
        return None, False, veh_loc, 'cancelled'
    if status == '' and not get("Driver's Name") and not get('Client'):
        return None, False, veh_loc, 'blank'
    if actual or status == 'DELIVERED' or veh_loc in DELIVERED_LOCATIONS:
        return 'Delivered', True, veh_loc, None
    if veh_loc in DEALERSHIP_LOCATIONS:
        return ('Ready for Pickup' if pdi_done else 'Pre-Delivery Inspection'), True, veh_loc, None
    if veh_loc in TRANSIT_LOCATIONS:
        return 'In Transit', False, veh_loc, None
    if status == 'NOT AVAILABLE' or veh_loc == 'NO ALLOCATION' or veh_loc == '':
        return 'Scheduled', False, veh_loc, None
    return 'Scheduled', False, veh_loc, None


def derive_contact_status(get, stage):
    if stage == 'Delivered':
        return 'Contacted'
    if (get('Follow Up Comments') or '').strip():
        return 'Contacted'
    if (get('Sales Comments') or '').strip():
        return 'Awaiting Reply'
    return 'Not Contacted'


def build_vehicle(get):
    parts = []
    veh_type = (get('VEH. Type') or '').strip()
    model = (get('VEH. Model') or '').strip()
    variant = (get('VEH. Variant') or '').strip()
    colour = (get('VEH. Colour') or '').strip()
    if veh_type:
        parts.append(f"BYD {veh_type.title()}")
    if model and model.lower() != veh_type.lower():
        parts.append(model.title())
    if variant:
        parts.append(variant.title())
    label = ' '.join(parts).strip()
    if colour:
        label = f"{label} \u00b7 {colour.title()}" if label else colour.title()
    return label or 'BYD Vehicle'


def split_addons(get):
    out = []
    for key in ('ACCESSORIES', 'A/Mkt External Accessories', 'Aftermarket Products'):
        v = get(key)
        if not v:
            continue
        for line in re.split(r'[\n;,]+', str(v)):
            line = line.strip(' -*')
            if line and len(line) > 2:
                out.append(line[:120])
    seen, deduped = set(), []
    for a in out:
        if a.lower() not in seen:
            seen.add(a.lower()); deduped.append(a)
    return deduped


def aftermarket_text(get):
    parts = []
    for key in ('A/M Comments', 'Stock Control Notes'):
        v = get(key)
        if v and str(v).strip():
            parts.append(f"[{key}] {str(v).strip()}")
    return '\n'.join(parts) or None


def comments_blocks(get, vehicle_at):
    out = []
    sources = [
        ('Pre-Delivery Comments', 'Pre-Delivery'),
        ('Sales Comments', 'Sales'),
        ('Follow Up Comments', 'Follow-up'),
    ]
    for col_name, label in sources:
        body = get(col_name)
        if not body:
            continue
        clean = str(body).strip()
        if not clean or clean.upper() == 'PDI COMPLETED':
            continue
        out.append({
            'id': uuid.uuid4().hex,
            'author_id': None,
            'author_name': f'Harmony Auto \u00b7 {label}',
            'body': clean,
            'created_at': datetime.now(timezone.utc),
        })
    if vehicle_at:
        out.append({
            'id': uuid.uuid4().hex,
            'author_id': None,
            'author_name': 'System',
            'body': f'Vehicle location at import: {vehicle_at}',
            'created_at': datetime.now(timezone.utc),
        })
    return out


def import_harmony(source, db, *, replace: bool = True, dry_run: bool = False) -> dict:
    """Import Harmony Auto xlsx records.

    Args:
        source: file path (str), bytes, or file-like object containing xlsx data
        db: pymongo sync database object
        replace: if True, delete previously imported harmony rows before inserting
        dry_run: if True, parse and categorise but DO NOT touch the database
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    summary = {
        'wiped': 0,
        'inserted': 0,
        'skipped_cancelled': 0,
        'skipped_blank': 0,
        'by_stage': {},
        'arrived': 0,
        'unassigned': 0,
        'auto_assigned': 0,
        'dry_run': dry_run,
        'replace': replace,
    }

    if replace and not dry_run:
        wiped = db.clients.delete_many({'imported_from': 'harmony-xlsx'})
        summary['wiped'] = wiped.deleted_count
    elif replace and dry_run:
        summary['wiped'] = db.clients.count_documents({'imported_from': 'harmony-xlsx'})

    user_map = {u.get('name', '').lower(): u for u in db.users.find({'active': True})}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue

        def get(k, _row=r):
            i = col(k)
            return _row[i] if i is not None and i < len(_row) else None

        name_raw = get("Driver's Name") or get('Client') or get('Company Name')
        if not name_raw:
            summary['skipped_blank'] += 1
            continue

        stage, arrived, vehicle_at, skip = derive_stage(get)
        if skip:
            summary[f'skipped_{skip}'] = summary.get(f'skipped_{skip}', 0) + 1
            continue

        order_no = (str(get('Order No.') or '').strip()) or None

        # If we did not wipe, dedupe by vy_order_id
        if not replace and order_no and db.clients.find_one({'vy_order_id': order_no}):
            continue

        phone = normalise_phone(get('Phone Number'))
        email = (str(get('Email') or '').strip().lower()) or None
        delivery_date = to_iso_date(get('Estimated Delivery')) or to_iso_date(get('Actual Delivery'))
        actual_delivery = to_iso_date(get('Actual Delivery'))
        order_date = to_iso_date(get('Order Date'))

        sales_person = (str(get('Sales Person') or '').strip()) or None
        delivery_consultant = (str(get('Delivery Consultant') or '').strip()) or None

        assigned_agent_id = None
        if delivery_consultant:
            u = user_map.get(delivery_consultant.lower())
            if u:
                assigned_agent_id = u['id']
                summary['auto_assigned'] += 1

        addons = split_addons(get)
        aftermarket = aftermarket_text(get)
        comments = comments_blocks(get, vehicle_at if not arrived else None)
        contact_status = derive_contact_status(get, stage)

        suburb = (str(get('Suburb') or '').strip().title()) or None
        location = None
        if suburb:
            location = suburb if 'VIC' in suburb.upper() else f"{suburb}, VIC"

        now = datetime.now(timezone.utc)
        last_contacted: Optional[datetime] = now if contact_status in ('Contacted', 'Awaiting Reply') else None
        if stage == 'Delivered' and actual_delivery:
            try:
                last_contacted = datetime.fromisoformat(f'{actual_delivery}T12:00:00+00:00')
            except Exception:
                last_contacted = now

        client_doc = {
            'id': uuid.uuid4().hex,
            'name': title_case_name(name_raw),
            'phone': phone or '',
            'email': email,
            'vehicle': build_vehicle(get),
            'rego': (str(get('Rego No.') or '').strip()) or None,
            'vin': (str(get('Vin No.') or '').strip()) or None,
            'delivery_date': delivery_date,
            'stage': stage,
            'salesperson': sales_person,
            'notes': None,
            'address': None,
            'location': location,
            'deal_type': 'Retail',
            'vy_order_id': order_no,
            'vy_stock_id': (str(get('Stock No.') or '').strip()) or None,
            'stripe_customer_id': None,
            'arrived': arrived,
            'arrived_at': now if arrived else None,
            'contact_status': contact_status,
            'last_contacted_at': last_contacted,
            'assigned_agent_id': assigned_agent_id,
            'accessories': [],
            'aftermarket_notes': aftermarket,
            'addons': addons,
            'imported_from': 'harmony-xlsx',
            'imported_at': now,
            'comments': comments,
            'created_at': datetime.fromisoformat(f'{order_date}T00:00:00+00:00') if order_date else now,
            'updated_at': now,
        }
        if not dry_run:
            db.clients.insert_one(client_doc)
        summary['inserted'] += 1
        summary['by_stage'][stage] = summary['by_stage'].get(stage, 0) + 1
        if arrived:
            summary['arrived'] += 1
        if not assigned_agent_id:
            summary['unassigned'] += 1

    return summary


def import_harmony_from_bytes(data: bytes, *, replace: bool = True, dry_run: bool = False) -> dict:
    """Open a sync MongoClient, run the import, return the summary.

    Used by the FastAPI upload endpoint via asyncio.to_thread.
    """
    mongo_url = os.environ['MONGO_URL']
    db_name = os.environ.get('DB_NAME', 'test_database')
    client = MongoClient(mongo_url)
    try:
        return import_harmony(data, client[db_name], replace=replace, dry_run=dry_run)
    finally:
        client.close()
